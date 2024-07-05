import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def df_in_xlsx(df: pd.DataFrame, filename: str, max_width: int = 50):
    """Запись датафрейма в файл Excel с условным форматированием"""
    workbook = Workbook()
    sheet = workbook.active

    # Записываем заголовки колонок
    sheet.append(df.columns.tolist())

    # Записываем данные из DataFrame
    for row in dataframe_to_rows(df, index=False, header=False):
        sheet.append(row)

    # Устанавливаем ширину колонок
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, max_width)
        sheet.column_dimensions[column].width = adjusted_width

    # Добавляем фильтры на колонки
    sheet.auto_filter.ref = sheet.dimensions

    # Условное форматирование: если значение в колонке "Проверено" равно "Нет", заливаем строку красным
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    for row in sheet.iter_rows(min_row=2, min_col=df.columns.get_loc("Проверено") + 1, max_col=df.columns.get_loc("Проверено") + 1):
        for cell in row:
            if cell.value == "Нет":
                for cell_in_row in sheet[cell.row]:
                    cell_in_row.fill = red_fill
                break

    workbook.save(f"{filename}.xlsx")
